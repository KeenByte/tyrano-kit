#!/usr/bin/env python3
"""
Machine translation of XLSX file via Google Translate or MyMemory.
Free, no API key required.

Install:
    pip install deep-translator openpyxl

Usage:
    python translate_xlsx.py translations.xlsx                              # Google EN->RU
    python translate_xlsx.py translations.xlsx --engine mymemory            # MyMemory EN->RU
    python translate_xlsx.py translations.xlsx --from ja --to en            # Google JA->EN
    python translate_xlsx.py translations.xlsx --engine google --to de      # Google EN->DE
    python translate_xlsx.py translations.xlsx --list                       # list languages
"""

import sys
import time
import os
import argparse

# --- Default settings ---
BATCH_SIZE = 10
DELAY_BETWEEN = 0.5
SAVE_EVERY = 50
SOURCE_COL = 5           # E = Original
TARGET_COL = 6           # F = Translation
HEADER_ROW = 1
DEFAULT_ENGINE = 'google'


# ============================================================
# Translation engines (all via deep-translator)
# ============================================================

class GoogleEngine:
    """Google Translate via deep-translator (free, no API key)."""
    name = 'Google Translate'

    def __init__(self, source_lang, target_lang):
        from deep_translator import GoogleTranslator
        self._cls = GoogleTranslator
        self.source = source_lang
        self.target = target_lang
        # Test that languages are valid
        self._cls(source=self.source, target=self.target)

    def translate(self, text):
        tr = self._cls(source=self.source, target=self.target).translate(text)
        return tr if tr else text

    @staticmethod
    def list_languages():
        from deep_translator import GoogleTranslator
        return GoogleTranslator().get_supported_languages(as_dict=True)


class MyMemoryEngine:
    """MyMemory Translate via deep-translator (free, no API key, 5000 chars/day)."""
    name = 'MyMemory'

    def __init__(self, source_lang, target_lang):
        from deep_translator import MyMemoryTranslator
        self._cls = MyMemoryTranslator
        self.source = source_lang
        self.target = target_lang
        self._cls(source=self.source, target=self.target)

    def translate(self, text):
        tr = self._cls(source=self.source, target=self.target).translate(text)
        return tr if tr else text

    @staticmethod
    def list_languages():
        from deep_translator import MyMemoryTranslator
        return MyMemoryTranslator().get_supported_languages(as_dict=True)


class DeepLEngine:
    """DeepL Translator -- best quality for Russian.
    Requires a free API key from https://www.deepl.com/pro-api
    Free tier: 500,000 chars/month.
    Set env var DEEPL_API_KEY or pass via --deepl-key."""
    name = 'DeepL'

    def __init__(self, source_lang, target_lang, api_key=None):
        self.api_key = api_key or os.environ.get("DEEPL_API_KEY", "")
        if not self.api_key:
            raise ValueError(
                "DeepL requires an API key.\n"
                "  Get a free key at: https://www.deepl.com/pro-api\n"
                "  Then either:\n"
                "    set DEEPL_API_KEY=your_key_here   (env var)\n"
                "    or use --deepl-key your_key_here   (command line)")
        from deep_translator import DeeplTranslator
        self._cls = DeeplTranslator
        # DeepL uses uppercase + variant codes (e.g. EN, RU, PT-BR)
        self.source = source_lang
        self.target = target_lang
        # Validate
        self._cls(api_key=self.api_key, source=self.source, target=self.target)

    def translate(self, text):
        tr = self._cls(
            api_key=self.api_key, source=self.source, target=self.target
        ).translate(text)
        return tr if tr else text

    @staticmethod
    def list_languages():
        return {
            'bg': 'Bulgarian',    'cs': 'Czech',        'da': 'Danish',
            'de': 'German',       'el': 'Greek',        'en': 'English',
            'es': 'Spanish',      'et': 'Estonian',      'fi': 'Finnish',
            'fr': 'French',       'hu': 'Hungarian',     'id': 'Indonesian',
            'it': 'Italian',      'ja': 'Japanese',      'ko': 'Korean',
            'lt': 'Lithuanian',   'lv': 'Latvian',       'nb': 'Norwegian',
            'nl': 'Dutch',        'pl': 'Polish',        'pt': 'Portuguese',
            'ro': 'Romanian',     'ru': 'Russian',       'sk': 'Slovak',
            'sl': 'Slovenian',    'sv': 'Swedish',       'tr': 'Turkish',
            'uk': 'Ukrainian',    'zh': 'Chinese',
        }


class LibreEngine:
    """LibreTranslate -- free, open source, no API key needed.
    Uses public mirror. Quality is decent, fully free."""
    name = 'LibreTranslate'

    def __init__(self, source_lang, target_lang):
        from deep_translator import LibreTranslator
        self._cls = LibreTranslator
        self.source = source_lang
        self.target = target_lang
        self._cls(source=self.source, target=self.target)

    def translate(self, text):
        tr = self._cls(source=self.source, target=self.target).translate(text)
        return tr if tr else text

    @staticmethod
    def list_languages():
        return {
            'en': 'English',    'ru': 'Russian',    'de': 'German',
            'fr': 'French',     'es': 'Spanish',    'it': 'Italian',
            'pt': 'Portuguese', 'ja': 'Japanese',   'ko': 'Korean',
            'zh': 'Chinese',    'ar': 'Arabic',     'tr': 'Turkish',
            'pl': 'Polish',     'nl': 'Dutch',      'uk': 'Ukrainian',
            'hi': 'Hindi',      'vi': 'Vietnamese', 'id': 'Indonesian',
            'cs': 'Czech',      'fi': 'Finnish',    'el': 'Greek',
            'he': 'Hebrew',     'hu': 'Hungarian',  'sv': 'Swedish',
        }


ENGINES = {
    'google': GoogleEngine,
    'deepl': DeepLEngine,
    'libre': LibreEngine,
    'mymemory': MyMemoryEngine,
}


# ============================================================
# Main logic
# ============================================================

def list_languages(engine_name):
    """Show available languages for the selected engine."""
    engine_cls = ENGINES.get(engine_name)
    if not engine_cls:
        print(f"Unknown engine: {engine_name}")
        print(f"Available: {', '.join(ENGINES.keys())}")
        return
    langs = engine_cls.list_languages()
    print(f"\nEngine: {engine_cls.name}")
    print(f"Available languages ({len(langs)}):\n")
    print(f"  {'Code':<8} {'Language':<25}")
    print(f"  {'-'*8} {'-'*25}")
    for key, val in sorted(langs.items()):
        print(f"  {key:<8} {val:<25}")
    print(f"\nExample: python translate_xlsx.py file.xlsx --engine {engine_name} --from en --to ru\n")


def translate_one(engine, text):
    """Translate a single string via the engine."""
    try:
        tr = engine.translate(text)
        return tr if tr else text
    except Exception as e:
        print(f"  [!] Error: '{text[:40]}...': {e}")
        return ""


def main():
    parser = argparse.ArgumentParser(
        description='Machine translation of XLSX (Google / DeepL / Libre / MyMemory)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Engines:
  google   -- Google Translate (free, no key, good for most languages)
  deepl    -- DeepL (best quality for RU/DE/etc, free key from deepl.com)
  libre    -- LibreTranslate (free, open source, decent quality)
  mymemory -- MyMemory (free, 5000 chars/day limit)

Examples:
  python translate_xlsx.py translations.xlsx                          # Google EN->RU
  python translate_xlsx.py translations.xlsx --engine deepl --deepl-key YOUR_KEY
  python translate_xlsx.py translations.xlsx --engine libre           # LibreTranslate
  python translate_xlsx.py translations.xlsx --engine mymemory        # MyMemory EN->RU
  python translate_xlsx.py translations.xlsx --from ja --to en        # Google JA->EN
  python translate_xlsx.py translations.xlsx --list                   # languages (Google)
  python translate_xlsx.py translations.xlsx --list --engine deepl    # languages (DeepL)
        """
    )
    parser.add_argument('xlsx', nargs='?', help='XLSX file to translate')
    parser.add_argument('--engine', default=DEFAULT_ENGINE,
                        choices=ENGINES.keys(),
                        help=f'Translation engine (default: {DEFAULT_ENGINE})')
    parser.add_argument('--from', dest='source_lang', default='en',
                        help='Source language code (default: en)')
    parser.add_argument('--to', dest='target_lang', default='ru',
                        help='Target language code (default: ru)')
    parser.add_argument('--deepl-key', default=None,
                        help='DeepL API key (or set DEEPL_API_KEY env var)')
    parser.add_argument('--list', action='store_true',
                        help='Show list of available languages')
    parser.add_argument('--batch-size', type=int, default=BATCH_SIZE,
                        help=f'Rows per batch (default: {BATCH_SIZE})')
    parser.add_argument('--delay', type=float, default=DELAY_BETWEEN,
                        help=f'Delay between batches in seconds (default: {DELAY_BETWEEN})')

    args = parser.parse_args()

    if args.list:
        list_languages(args.engine)
        return

    if not args.xlsx:
        parser.print_help()
        sys.exit(1)

    xlsx_path = args.xlsx
    source_lang = args.source_lang
    target_lang = args.target_lang
    engine_name = args.engine
    batch_size = args.batch_size
    delay = args.delay

    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    # Create engine
    engine_cls = ENGINES[engine_name]
    try:
        if engine_name == 'deepl':
            engine = engine_cls(source_lang, target_lang, api_key=args.deepl_key)
        else:
            engine = engine_cls(source_lang, target_lang)
    except ImportError:
        print("Missing dependency: deep-translator")
        print("Install: pip install deep-translator")
        sys.exit(1)
    except ValueError as e:
        print(str(e))
        sys.exit(1)

    print(f"Engine: {engine_cls.name} ({source_lang} -> {target_lang})")

    # Load XLSX
    from openpyxl import load_workbook

    base, ext = os.path.splitext(xlsx_path)
    output_path = f"{base}_translated{ext}"

    if os.path.exists(output_path):
        print(f"Found existing {output_path}, resuming translation...")
        wb = load_workbook(output_path)
    else:
        print(f"Loading {xlsx_path}...")
        wb = load_workbook(xlsx_path)

    ws = wb.active
    total_rows = ws.max_row - HEADER_ROW

    # Count
    skipped = 0
    to_translate = []
    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        original = ws.cell(row=row_idx, column=SOURCE_COL).value
        translation = ws.cell(row=row_idx, column=TARGET_COL).value
        if translation and str(translation).strip():
            skipped += 1
        elif original and str(original).strip():
            to_translate.append(row_idx)

    print(f"Total rows: {total_rows}")
    print(f"Already translated: {skipped}")
    print(f"Remaining: {len(to_translate)}")
    print(f"Output file: {output_path}")
    print("-" * 50)

    if not to_translate:
        print("Nothing to translate -- all done!")
        return

    translated_count = 0
    last_save = 0
    pause_file = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              ".pause_signal")

    try:
        for i in range(0, len(to_translate), batch_size):
            # Check for pause signal between batches
            while os.path.exists(pause_file):
                time.sleep(0.5)

            batch_rows = to_translate[i:i + batch_size]

            for row_idx in batch_rows:
                text = str(ws.cell(row=row_idx, column=SOURCE_COL).value).strip()
                tr_text = translate_one(engine, text)
                ws.cell(row=row_idx, column=TARGET_COL, value=tr_text)
                translated_count += 1

            done = skipped + translated_count
            pct = done / total_rows * 100
            sample = str(ws.cell(row=batch_rows[0], column=SOURCE_COL).value)[:50]
            print(f"  [{done}/{total_rows}] {pct:.1f}%  last: {sample}...")

            if translated_count - last_save >= SAVE_EVERY:
                wb.save(output_path)
                last_save = translated_count
                print(f"  Saved ({output_path})")

            time.sleep(delay)

    except KeyboardInterrupt:
        print("\n\nInterrupted by user (Ctrl+C)")

    wb.save(output_path)
    print(f"\n{'=' * 50}")
    print(f"Done! Translated in this run: {translated_count}")
    print(f"File saved: {output_path}")


if __name__ == '__main__':
    main()
