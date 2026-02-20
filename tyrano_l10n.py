#!/usr/bin/env python3
"""
TyranoScript Localization Tool
Extracts translatable strings from .ks files → XLSX
Applies translations from XLSX → translated .ks files
"""

import re
import os
import sys
import glob
import argparse
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Patterns ---
TAG_RE = re.compile(r'\[.*?\]')
GLINK_TEXT_RE = re.compile(r'\[glink\b[^\]]*\btext="([^"]*)"[^\]]*\]', re.IGNORECASE)
CHAR_NAME_RE = re.compile(r'^#(.+)$')
EMB_RE = re.compile(r'\[emb\s+exp="[^"]*"\]')

def is_tag_only(line):
    """Check if line is purely tags/commands with no translatable text."""
    stripped = TAG_RE.sub('', line).strip()
    return len(stripped) == 0

def extract_strings(ks_path, rel_path=None):
    """Extract translatable strings from a .ks file."""
    entries = []
    filename = rel_path.replace('\\', '/') if rel_path else os.path.basename(ks_path)

    with open(ks_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    in_text_block = False
    text_block_start = None

    for i, raw_line in enumerate(lines, 1):
        line = raw_line.rstrip('\r\n')

        # Detect text block boundaries
        if '[tb_start_text' in line:
            in_text_block = True
            text_block_start = i
            continue
        if '[_tb_end_text]' in line:
            in_text_block = False
            continue

        # Inside text block: extract narrative/dialogue
        if in_text_block:
            # Character name line
            m_char = CHAR_NAME_RE.match(line.strip())
            if m_char and m_char.group(1).strip():
                name = m_char.group(1).strip()
                entries.append({
                    'file': filename,
                    'line': i,
                    'type': 'character_name',
                    'original': name,
                    'context': f'Text block starting line {text_block_start}',
                })
                continue

            # Empty line or bare #
            if line.strip() == '' or line.strip() == '#':
                continue

            # Tag-only line (like [p], [cm], etc.)
            if is_tag_only(line.strip()):
                continue

            # Text line (may contain inline tags like [p], [emb ...])
            text = line.strip()
            # Remove trailing [p] for cleaner display but keep original
            clean = TAG_RE.sub('', text).strip()
            if clean:
                entries.append({
                    'file': filename,
                    'line': i,
                    'type': 'dialogue',
                    'original': clean,
                    'context': f'Text block starting line {text_block_start}',
                    'raw_line': text,
                })

        # Outside text blocks: extract glink button text
        glink_matches = GLINK_TEXT_RE.findall(line)
        for gtext in glink_matches:
            clean_text = gtext.replace('&nbsp;', ' ')
            if clean_text.strip():
                entries.append({
                    'file': filename,
                    'line': i,
                    'type': 'button',
                    'original': clean_text,
                    'context': 'Menu button / glink',
                    'raw_glink_text': gtext,
                })

    return entries


def extract_to_xlsx(input_dir, output_xlsx):
    """Extract all .ks files from input_dir → XLSX translation table."""
    ks_files = sorted(glob.glob(os.path.join(input_dir, '**', '*.ks'), recursive=True))
    if not ks_files:
        print(f"No .ks files found in {input_dir}")
        return

    all_entries = []
    for ks in ks_files:
        rel_path = os.path.relpath(ks, input_dir)
        entries = extract_strings(ks, rel_path)
        all_entries.extend(entries)

    print(f"Found {len(all_entries)} translatable strings in {len(ks_files)} files")

    # Create XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = ["ID", "File", "Line", "Type", "Original", "Translation", "Context"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Type colors
    type_fills = {
        'dialogue': PatternFill("solid", fgColor="FFFFFF"),
        'character_name': PatternFill("solid", fgColor="FFF2CC"),
        'button': PatternFill("solid", fgColor="D6E4F0"),
    }

    # Data
    for idx, e in enumerate(all_entries, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=e['file']).border = thin_border
        ws.cell(row=row, column=3, value=e['line']).border = thin_border
        ws.cell(row=row, column=4, value=e['type']).border = thin_border

        orig_cell = ws.cell(row=row, column=5, value=e['original'])
        orig_cell.border = thin_border
        orig_cell.alignment = Alignment(wrap_text=True)

        trans_cell = ws.cell(row=row, column=6, value='')
        trans_cell.border = thin_border
        trans_cell.alignment = Alignment(wrap_text=True)
        trans_cell.font = Font(color="0000FF", name="Arial")  # Blue = input

        ctx_cell = ws.cell(row=row, column=7, value=e.get('context', ''))
        ctx_cell.border = thin_border
        ctx_cell.font = Font(color="808080", italic=True, name="Arial")

        fill = type_fills.get(e['type'], PatternFill())
        for c in range(1, 8):
            if e['type'] in type_fills:
                ws.cell(row=row, column=c).fill = fill

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 7
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 30

    # Freeze header row
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_xlsx)
    print(f"Saved: {output_xlsx}")


def apply_translations(input_dir, xlsx_path, output_dir):
    """Read translations from XLSX and apply to .ks files."""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Build translation map: (file, line, type) → translation
    translations = {}
    # Also build button text map: (file, original_text) → translation (for glink)
    button_map = {}

    for row in ws.iter_rows(min_row=2, values_only=False):
        cells = [c.value for c in row]
        if len(cells) < 6:
            continue
        _id, filename, line_num, str_type, original, translation = cells[:6]
        if not translation or not str(translation).strip():
            continue

        key = (str(filename).replace('\\', '/'), int(line_num), str(str_type))
        translations[key] = str(translation).strip()

        if str_type == 'button':
            btn_key = (str(filename).replace('\\', '/'), str(original).strip())
            button_map[btn_key] = str(translation).strip()

    print(f"Loaded {len(translations)} translations")

    os.makedirs(output_dir, exist_ok=True)

    ks_files = sorted(glob.glob(os.path.join(input_dir, '**', '*.ks'), recursive=True))

    for ks_path in ks_files:
        rel_path = os.path.relpath(ks_path, input_dir)
        filename = rel_path.replace('\\', '/')

        with open(ks_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        new_lines = []
        in_text_block = False

        for i, raw_line in enumerate(lines, 1):
            line = raw_line

            # Track text blocks
            if '[tb_start_text' in line:
                in_text_block = True
                new_lines.append(line)
                continue
            if '[_tb_end_text]' in line:
                in_text_block = False
                new_lines.append(line)
                continue

            if in_text_block:
                stripped = line.rstrip('\r\n').strip()

                # Character name
                m_char = CHAR_NAME_RE.match(stripped)
                if m_char and m_char.group(1).strip():
                    key = (filename, i, 'character_name')
                    if key in translations:
                        new_lines.append(f'#{translations[key]}\n')
                        continue

                # Dialogue line
                key = (filename, i, 'dialogue')
                if key in translations and not is_tag_only(stripped) and stripped != '#' and stripped != '':
                    tr = translations[key]
                    # Preserve inline tags: keep [p] at end, [emb ...] etc.
                    tags_at_end = ''
                    if stripped.endswith('[p]'):
                        tags_at_end = '[p]'
                    # Reconstruct: preserve embedded tags
                    emb_tags = EMB_RE.findall(stripped)
                    new_text = tr
                    for emb in emb_tags:
                        if emb not in new_text:
                            new_text = new_text + emb
                    if tags_at_end and not new_text.endswith('[p]'):
                        new_text = new_text + '[p]'
                    new_lines.append(new_text + '\n')
                    continue

            # glink button text replacement
            if '[glink' in line.lower():
                glink_matches = GLINK_TEXT_RE.finditer(line)
                modified_line = line
                for m in glink_matches:
                    original_gtext = m.group(1)
                    clean_original = original_gtext.replace('&nbsp;', ' ')
                    btn_key = (filename, clean_original.strip())
                    if btn_key in button_map:
                        new_gtext = button_map[btn_key].replace(' ', '&nbsp;')
                        modified_line = modified_line.replace(
                            f'text="{original_gtext}"',
                            f'text="{new_gtext}"'
                        )
                line = modified_line

            new_lines.append(line)

        out_path = os.path.join(output_dir, rel_path)
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        with open(out_path, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)

    print(f"Translated files saved to: {output_dir}")


def main():
    parser = argparse.ArgumentParser(description='TyranoScript Localization Tool')
    sub = parser.add_subparsers(dest='command')

    # Extract command
    p_ext = sub.add_parser('extract', help='Extract strings from .ks → XLSX')
    p_ext.add_argument('input_dir', help='Directory with .ks files')
    p_ext.add_argument('-o', '--output', default='translations.xlsx', help='Output XLSX path')

    # Apply command
    p_app = sub.add_parser('apply', help='Apply translations from XLSX → .ks files')
    p_app.add_argument('input_dir', help='Directory with original .ks files')
    p_app.add_argument('xlsx', help='XLSX file with translations')
    p_app.add_argument('-o', '--output', default='translated', help='Output directory')

    args = parser.parse_args()

    if args.command == 'extract':
        extract_to_xlsx(args.input_dir, args.output)
    elif args.command == 'apply':
        apply_translations(args.input_dir, args.xlsx, args.output)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
